"""
exports/excel.py

OpenPyXL formatted workbook generation for Daylite Person Quality findings and summaries.
"""

import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def generate_xlsx(sheet_title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
    """
    Builds a styled Excel (.xlsx) workbook. Uses write_only streaming mode for datasets > 1,000 rows
    to minimize memory consumption, and caps auto column widths at max 60 chars.
    """
    if len(rows) > 1000:
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet(title=sheet_title[:31])
        ws.append(headers)
        for r in rows:
            ws.append(r)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Calibri", size=10)
    data_align = Alignment(vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 26

    for row_idx, row_data in enumerate(rows, start=2):
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 20
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
