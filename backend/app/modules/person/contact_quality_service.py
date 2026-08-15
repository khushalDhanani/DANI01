import csv
import io
import logging
import time
from datetime import datetime, timezone
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.db.mssql import execute_readonly_query
from app.modules.person.contact_quality_schemas import (
    ContactQualityIssueItem,
    ContactQualityIssuesResponse,
    ContactQualityIssueType,
    ContactQualitySummaryResponse,
)

logger = logging.getLogger(__name__)


def generate_csv(headers: list[str], rows: list[list[Any]]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for r in rows:
        writer.writerow(r)
    return output.getvalue().encode("utf-8")


def generate_xlsx(sheet_title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Calibri", size=10)
    data_align = Alignment(vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 26

    for row_idx, row_data in enumerate(rows, start=2):
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 20
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



def mask_contact_value(contact_type: str, raw_value: str | None) -> str:
    """
    Returns the exact raw contact value without masking so quality anomalies are fully visible.
    """
    if not raw_value or not raw_value.strip():
        return "—"
    return raw_value.strip()


def _resolve_order_clause(
    sort_by: str = "PersonID",
    sort_order: str = "desc",
    default_clause: str = "p.PersonID DESC",
) -> str:
    """
    Safely resolves sorting column and direction to prevent SQL injection.
    """
    direction = "ASC" if sort_order.lower().strip() == "asc" else "DESC"
    col = sort_by.lower().strip()

    if col in ("personid", "id"):
        return f"p.PersonID {direction}"
    elif col in ("personname", "name"):
        return f"ISNULL(p.PersonFirstName, '') {direction}, ISNULL(p.PersonLastName, '') {direction}"
    elif col in ("currentvalue", "value"):
        return f"CurrentValue {direction}, p.PersonID {direction}"
    elif col in ("severity",):
        return f"Severity {direction}, p.PersonID {direction}"
    return default_clause


PERSON_NAME_SQL = """
CASE 
    WHEN NULLIF(LTRIM(RTRIM(ISNULL(p.PersonFirstName, '') + ' ' + ISNULL(p.PersonLastName, ''))), '') IS NOT NULL 
    THEN LTRIM(RTRIM(ISNULL(p.PersonFirstName, '') + ' ' + ISNULL(p.PersonLastName, '')))
    ELSE 'Person #' + CAST(p.PersonID AS varchar)
END
""".strip()

QUALIFYING_EMAIL_EXISTS_SQL = """
EXISTS (
    SELECT 1 
    FROM dbo.DLPersonPhoneEmailURLDet c 
    LEFT JOIN dbo.DLLabelTypeMst l ON c.LabelTypeID = l.LabelTypeID
    WHERE c.PersionID = p.PersonID 
      AND (l.LabelType = 'EMail' OR (c.LabelTypeID IS NULL AND c.TypeValue LIKE '%@%'))
      AND c.TypeValue IS NOT NULL 
      AND LTRIM(RTRIM(c.TypeValue)) <> ''
      AND ISNULL(c.PersonPhoneIsActive, 1) = 1
)
""".strip()

QUALIFYING_PHONE_EXISTS_SQL = """
EXISTS (
    SELECT 1 
    FROM dbo.DLPersonPhoneEmailURLDet c 
    LEFT JOIN dbo.DLLabelTypeMst l ON c.LabelTypeID = l.LabelTypeID
    WHERE c.PersionID = p.PersonID 
      AND (l.LabelType = 'PhoneNumbers' OR (c.LabelTypeID IS NULL AND c.TypeValue NOT LIKE '%@%' AND c.TypeValue NOT LIKE 'http%' AND c.TypeValue NOT LIKE 'www%'))
      AND c.TypeValue IS NOT NULL 
      AND LTRIM(RTRIM(c.TypeValue)) <> ''
      AND ISNULL(c.PersonPhoneIsActive, 1) = 1
)
""".strip()


class ContactQualityService:
    """
    Comprehensive Service for analyzing Person data quality across contacts, addresses,
    identity chronology, employment consistency, compliance, links, and audit trails.
    """

    async def get_contact_quality_summary(self) -> ContactQualitySummaryResponse:
        start_time = time.perf_counter()

        sql = f"""
        WITH ClassifiedContacts AS (
            SELECT 
                c.PersonPhoneID,
                c.PersionID AS PersonID,
                c.LabelTypeID,
                l.LableValue AS LabelName,
                c.TypeValue,
                c.IsVerified,
                c.IsPrimary,
                c.PersonPhoneIsActive,
                CASE 
                    WHEN l.LabelType = 'EMail' OR (c.LabelTypeID IS NULL AND c.TypeValue LIKE '%@%') THEN 'EMAIL'
                    WHEN l.LabelType = 'PhoneNumbers' OR (c.LabelTypeID IS NULL AND c.TypeValue NOT LIKE '%@%' AND c.TypeValue NOT LIKE 'http%' AND c.TypeValue NOT LIKE 'www%') THEN 'PHONE'
                    WHEN l.LabelType = 'URL' OR (c.LabelTypeID IS NULL AND (c.TypeValue LIKE 'http%' OR c.TypeValue LIKE 'www%')) THEN 'URL'
                    ELSE 'OTHER'
                END AS ContactCategory,
                LOWER(LTRIM(RTRIM(c.TypeValue))) AS NormalizedEmail,
                REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(LTRIM(RTRIM(c.TypeValue)), ' ', ''), '-', ''), '+', ''), '(', ''), ')', ''), '.', ''), '/', '') AS NormalizedPhone
            FROM dbo.DLPersonPhoneEmailURLDet c
            JOIN dbo.DLPersonMst p ON p.PersonID = c.PersionID
            LEFT JOIN dbo.DLLabelTypeMst l ON c.LabelTypeID = l.LabelTypeID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0
        )
        SELECT 
            -- 1. Contact Quality (Valid Active Persons Only)
            (SELECT COUNT_BIG(1) 
             FROM dbo.DLPersonMst p 
             WHERE p.PersonIsActive = 1 
               AND ISNULL(p.PersonIsDeleted, 0) = 0 
               AND NOT {QUALIFYING_EMAIL_EXISTS_SQL}
            ) AS persons_without_email,
            (SELECT COUNT_BIG(1) 
             FROM dbo.DLPersonMst p 
             WHERE p.PersonIsActive = 1 
               AND ISNULL(p.PersonIsDeleted, 0) = 0 
               AND NOT {QUALIFYING_PHONE_EXISTS_SQL}
            ) AS persons_without_phone,
            (SELECT COUNT_BIG(1) FROM ClassifiedContacts WHERE ContactCategory = 'EMAIL' AND (
                TypeValue IS NULL OR LTRIM(RTRIM(TypeValue)) = '' 
                OR TypeValue NOT LIKE '%_@__%.__%'
                OR TypeValue LIKE '% %'
                OR TypeValue LIKE '%@%@%'
                OR TypeValue LIKE '%.@%'
                OR TypeValue LIKE '%@.%'
            )) AS invalid_emails,
            (SELECT COUNT_BIG(1) FROM ClassifiedContacts WHERE ContactCategory = 'PHONE' AND (
                TypeValue IS NULL OR LTRIM(RTRIM(TypeValue)) = ''
                OR (
                    CASE 
                        WHEN LOWER(ISNULL(LabelName, '')) LIKE '%ext%' THEN
                            CASE WHEN LEN(NormalizedPhone) = 4 AND NormalizedPhone NOT LIKE '%[^0-9]%' THEN 0 ELSE 1 END
                        ELSE
                            CASE WHEN LEN(NormalizedPhone) < 7 OR LEN(NormalizedPhone) > 15 OR NormalizedPhone LIKE '%[^0-9]%' OR NormalizedPhone IN ('0000000000', '1234567890', '9999999999', '1111111111') THEN 1 ELSE 0 END
                    END = 1
                )
            )) AS invalid_phones,
            (SELECT COUNT_BIG(1) FROM ClassifiedContacts WHERE ContactCategory = 'URL' AND (
                TypeValue IS NULL OR LTRIM(RTRIM(TypeValue)) = ''
                OR (TypeValue NOT LIKE 'http://%' AND TypeValue NOT LIKE 'https://%' AND TypeValue NOT LIKE 'www.%')
            )) AS invalid_urls,
            (SELECT COUNT_BIG(1) FROM ClassifiedContacts WHERE IsVerified = 0 OR IsVerified IS NULL) AS unverified_contacts,
            (SELECT COUNT_BIG(1) FROM (
                SELECT NormalizedEmail 
                FROM ClassifiedContacts 
                WHERE ContactCategory = 'EMAIL' AND NormalizedEmail <> '' AND NormalizedEmail IS NOT NULL
                GROUP BY NormalizedEmail 
                HAVING COUNT(DISTINCT PersonID) > 1
            ) t) AS duplicate_email_cross_persons,
            (SELECT COUNT_BIG(1) FROM (
                SELECT PersonID, NormalizedEmail 
                FROM ClassifiedContacts 
                WHERE ContactCategory = 'EMAIL' AND NormalizedEmail <> '' AND NormalizedEmail IS NOT NULL
                GROUP BY PersonID, NormalizedEmail 
                HAVING COUNT_BIG(1) > 1
            ) t) AS duplicate_email_same_person,
            (SELECT COUNT_BIG(1) FROM (
                SELECT NormalizedPhone 
                FROM ClassifiedContacts 
                WHERE ContactCategory = 'PHONE' AND NormalizedPhone <> '' AND NormalizedPhone IS NOT NULL AND LEN(NormalizedPhone) >= 7
                GROUP BY NormalizedPhone 
                HAVING COUNT(DISTINCT PersonID) > 1
            ) t) AS duplicate_phone_cross_persons,
            (SELECT COUNT_BIG(1) FROM (
                SELECT PersonID, NormalizedPhone 
                FROM ClassifiedContacts 
                WHERE ContactCategory = 'PHONE' AND NormalizedPhone <> '' AND NormalizedPhone IS NOT NULL AND LEN(NormalizedPhone) >= 7
                GROUP BY PersonID, NormalizedPhone 
                HAVING COUNT_BIG(1) > 1
            ) t) AS duplicate_phone_same_person,
            (SELECT COUNT_BIG(1) FROM (
                SELECT PersonID 
                FROM ClassifiedContacts 
                WHERE IsPrimary = 1 
                GROUP BY PersonID 
                HAVING COUNT_BIG(1) > 1
            ) t) AS persons_multiple_primary,
            (SELECT COUNT_BIG(1) FROM ClassifiedContacts WHERE IsPrimary = 1 AND (PersonPhoneIsActive = 0 OR PersonPhoneIsActive IS NULL)) AS primary_contact_inactive,

            -- 2. Address & Location Quality (Valid Active Persons Only)
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonAddressDet a JOIN dbo.DLPersonMst p ON p.PersonID = a.PersonID WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (a.PostalCode IS NULL OR LTRIM(RTRIM(a.PostalCode)) = '')) AS addr_missing_postal_code,
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonAddressDet a JOIN dbo.DLPersonMst p ON p.PersonID = a.PersonID WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND a.PostalCode IS NOT NULL AND LTRIM(RTRIM(a.PostalCode)) <> '' AND (LEN(LTRIM(RTRIM(a.PostalCode))) NOT IN (6, 5) OR LTRIM(RTRIM(a.PostalCode)) LIKE '%[^0-9]%')) AS addr_invalid_pin_format,
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonAddressDet a JOIN dbo.DLPersonMst p ON p.PersonID = a.PersonID WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (a.CityName IS NULL OR LTRIM(RTRIM(a.CityName)) = '') AND a.Street IS NOT NULL AND LTRIM(RTRIM(a.Street)) <> '') AS addr_street_without_city,
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonAddressDet a JOIN dbo.DLPersonMst p ON p.PersonID = a.PersonID WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND a.CityName IS NOT NULL AND LTRIM(RTRIM(a.CityName)) <> '' AND (a.StateName IS NULL OR LTRIM(RTRIM(a.StateName)) = '')) AS addr_city_without_state,
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonAddressDet a JOIN dbo.DLPersonMst p ON p.PersonID = a.PersonID WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (a.Latitude IS NULL OR a.Longitude IS NULL OR (a.Latitude = 0 AND a.Longitude = 0))) AS addr_missing_geocodes,
            (SELECT COUNT_BIG(1) FROM (
                SELECT a.PersonID, LOWER(LTRIM(RTRIM(a.Street))) AS s, LOWER(LTRIM(RTRIM(a.CityName))) AS c 
                FROM dbo.DLPersonAddressDet a
                JOIN dbo.DLPersonMst p ON p.PersonID = a.PersonID
                WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND a.Street IS NOT NULL AND LTRIM(RTRIM(a.Street)) <> '' 
                GROUP BY a.PersonID, LOWER(LTRIM(RTRIM(a.Street))), LOWER(LTRIM(RTRIM(a.CityName))) 
                HAVING COUNT_BIG(1) > 1
            ) t) AS addr_duplicate_same_person,

            -- 3. Profile & Chronological Integrity (Valid Active Persons Only)
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonMst WHERE PersonIsActive = 1 AND ISNULL(PersonIsDeleted, 0) = 0 AND PersonAnneversaryDate < PersonBirthDate) AS person_anniversary_before_birth,
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonMst WHERE PersonIsActive = 1 AND ISNULL(PersonIsDeleted, 0) = 0 AND (PersonBirthDate > GETDATE() OR PersonBirthDate < '1900-01-01')) AS person_invalid_birth_date,
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonMst WHERE PersonIsActive = 1 AND ISNULL(PersonIsDeleted, 0) = 0 AND PersonBirthDate IS NOT NULL AND (DATEDIFF(year, PersonBirthDate, GETDATE()) < 0 OR DATEDIFF(year, PersonBirthDate, GETDATE()) > 100)) AS person_birth_date_ancient,
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonMst WHERE PersonIsActive = 1 AND ISNULL(PersonIsDeleted, 0) = 0 AND (LOWER(PersonFirstName) IN ('test', 'admin', 'dummy', 'asdf', 'xyz', 'na', 'n/a', 'none', 'null', 'temp') OR LOWER(PersonLastName) IN ('test', 'admin', 'dummy', 'asdf', 'xyz', 'na', 'n/a', 'none', 'null', 'temp'))) AS person_suspicious_dummy_names,
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonMst WHERE PersonIsActive = 1 AND ISNULL(PersonIsDeleted, 0) = 0 AND (PersonLastName IS NULL OR LTRIM(RTRIM(PersonLastName)) = '') AND PersonFirstName IS NOT NULL AND LTRIM(RTRIM(PersonFirstName)) <> '') AS person_missing_lastname_only,

            -- 4. Employment & Lifecycle Consistency
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonMst WHERE EmpID IS NOT NULL AND PersonIsActive = 1 AND ISNULL(PersonIsDeleted, 0) = 0 AND (PersonTitle IS NULL OR LTRIM(RTRIM(PersonTitle)) = '')) AS active_emp_missing_title,
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonMst WHERE (PersonIsActive = 0 OR PersonIsActive IS NULL) AND EmpID IS NOT NULL) AS inactive_with_empid,
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonMst WHERE PersonIsActive = 1 AND PersonIsDeleted = 1) AS status_active_and_deleted,
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonMst WHERE PersonIsActive = 1 AND ISNULL(PersonIsDeleted, 0) = 0 AND PersonIsTemp = 1 AND PersonEntDt < DATEADD(day, -90, GETDATE())) AS stale_temp_persons,

            -- 5. Governance & Blacklist Compliance (Valid Active Persons Only)
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonMst WHERE PersonIsActive = 1 AND ISNULL(PersonIsDeleted, 0) = 0 AND PersonIsBlackList = 1 AND (PersonBlackListHODApprove = 0 OR PersonBlackListHODApprove IS NULL)) AS blacklist_unapproved,
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonMst WHERE PersonIsActive = 1 AND ISNULL(PersonIsDeleted, 0) = 0 AND PersonIsBlackList = 1 AND (PersonBlackListDate IS NULL OR PersonBlackListType IS NULL)) AS blacklist_missing_details,

            -- 6. Entity Linkages & Child Records (Valid Active Persons Only)
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonCompanyLinkDet l JOIN dbo.DLPersonMst p ON p.PersonID = l.PersonID WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND NOT EXISTS (SELECT 1 FROM dbo.DLCompanyMst c WHERE c.DLCompID = l.DLCompID)) AS company_orphan_links,
            (SELECT COUNT_BIG(1) FROM (
                SELECT l.PersonID, l.DLCompID 
                FROM dbo.DLPersonCompanyLinkDet l
                JOIN dbo.DLPersonMst p ON p.PersonID = l.PersonID
                WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND l.DLCompID IS NOT NULL 
                GROUP BY l.PersonID, l.DLCompID 
                HAVING COUNT_BIG(1) > 1
            ) t) AS company_duplicate_links,
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonCompanyLinkDet l JOIN dbo.DLPersonMst p ON p.PersonID = l.PersonID WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (l.CompPersonRoleID IS NULL OR l.CompPersonRoleID = 0)) AS company_missing_role,
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonExtraFieldValueDet e JOIN dbo.DLPersonMst p ON p.PersonID = e.PersonID WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND NOT EXISTS (SELECT 1 FROM dbo.DLExtraFieldDet m WHERE m.ExtraFieldID = e.ExtraFieldID)) AS extra_field_orphan_id,
            (SELECT COUNT_BIG(1) FROM (
                SELECT e.PersonID, e.ExtraFieldID 
                FROM dbo.DLPersonExtraFieldValueDet e
                JOIN dbo.DLPersonMst p ON p.PersonID = e.PersonID
                WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0
                GROUP BY e.PersonID, e.ExtraFieldID 
                HAVING COUNT_BIG(1) > 1
            ) t) AS extra_field_duplicate_entries,

            -- 7. Audit Trail & Sync Integration
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonMst WHERE PersonIsDeleted = 1 AND PersonDelDt IS NULL) AS deleted_missing_del_date,
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonMst WHERE PersonDelDt IS NOT NULL AND PersonEntDt IS NOT NULL AND PersonDelDt < PersonEntDt) AS audit_del_before_ent,
            (SELECT COUNT_BIG(1) FROM dbo.DLPersonMst WHERE PersonIsActive = 1 AND ISNULL(PersonIsDeleted, 0) = 0 AND IsContactSync = 1 AND ZimbraContactID IS NULL) AS sync_zimbra_missing_id,

            -- Scope: Valid Active Person Entity (dbo.DLPersonMst WHERE PersonIsActive = 1 AND ISNULL(PersonIsDeleted, 0) = 0)
            (SELECT COUNT_BIG(DISTINCT PersonID) FROM dbo.DLPersonMst WHERE PersonIsActive = 1 AND ISNULL(PersonIsDeleted, 0) = 0) AS total_persons_evaluated,
            (SELECT COUNT_BIG(DISTINCT PersonID) FROM dbo.DLPersonMst WHERE PersonIsActive = 0 OR PersonIsActive IS NULL) AS total_inactive_persons,
            (SELECT COUNT_BIG(DISTINCT PersonID) FROM dbo.DLPersonMst WHERE PersonIsDeleted = 1) AS total_deleted_persons
        ;
        """

        rows = execute_readonly_query(sql)
        row = rows[0] if rows else {}

        duration_ms = round((time.perf_counter() - start_time) * 1000, 2)
        calculated_at = datetime.now(timezone.utc).isoformat()

        return ContactQualitySummaryResponse(
            persons_without_email=int(row.get("persons_without_email") or 0),
            persons_without_phone=int(row.get("persons_without_phone") or 0),
            invalid_emails=int(row.get("invalid_emails") or 0),
            invalid_phones=int(row.get("invalid_phones") or 0),
            invalid_urls=int(row.get("invalid_urls") or 0),
            unverified_contacts=int(row.get("unverified_contacts") or 0),
            duplicate_email_cross_persons=int(row.get("duplicate_email_cross_persons") or 0),
            duplicate_email_same_person=int(row.get("duplicate_email_same_person") or 0),
            duplicate_phone_cross_persons=int(row.get("duplicate_phone_cross_persons") or 0),
            duplicate_phone_same_person=int(row.get("duplicate_phone_same_person") or 0),
            persons_multiple_primary=int(row.get("persons_multiple_primary") or 0),
            primary_contact_inactive=int(row.get("primary_contact_inactive") or 0),
            addr_missing_postal_code=int(row.get("addr_missing_postal_code") or 0),
            addr_invalid_pin_format=int(row.get("addr_invalid_pin_format") or 0),
            addr_street_without_city=int(row.get("addr_street_without_city") or 0),
            addr_city_without_state=int(row.get("addr_city_without_state") or 0),
            addr_missing_geocodes=int(row.get("addr_missing_geocodes") or 0),
            addr_duplicate_same_person=int(row.get("addr_duplicate_same_person") or 0),
            person_anniversary_before_birth=int(row.get("person_anniversary_before_birth") or 0),
            person_invalid_birth_date=int(row.get("person_invalid_birth_date") or 0),
            person_birth_date_ancient=int(row.get("person_birth_date_ancient") or 0),
            person_suspicious_dummy_names=int(row.get("person_suspicious_dummy_names") or 0),
            person_missing_lastname_only=int(row.get("person_missing_lastname_only") or 0),
            active_emp_missing_title=int(row.get("active_emp_missing_title") or 0),
            inactive_with_empid=int(row.get("inactive_with_empid") or 0),
            status_active_and_deleted=int(row.get("status_active_and_deleted") or 0),
            stale_temp_persons=int(row.get("stale_temp_persons") or 0),
            blacklist_unapproved=int(row.get("blacklist_unapproved") or 0),
            blacklist_missing_details=int(row.get("blacklist_missing_details") or 0),
            company_orphan_links=int(row.get("company_orphan_links") or 0),
            company_duplicate_links=int(row.get("company_duplicate_links") or 0),
            company_missing_role=int(row.get("company_missing_role") or 0),
            extra_field_orphan_id=int(row.get("extra_field_orphan_id") or 0),
            extra_field_duplicate_entries=int(row.get("extra_field_duplicate_entries") or 0),
            deleted_missing_del_date=int(row.get("deleted_missing_del_date") or 0),
            audit_del_before_ent=int(row.get("audit_del_before_ent") or 0),
            sync_zimbra_missing_id=int(row.get("sync_zimbra_missing_id") or 0),
            total_persons_evaluated=int(row.get("total_persons_evaluated") or 0),
            total_inactive_persons=int(row.get("total_inactive_persons") or 0),
            total_deleted_persons=int(row.get("total_deleted_persons") or 0),
            related_tables_checked=8,
            calculated_at=calculated_at,
            duration_ms=duration_ms,
        )

    async def get_contact_quality_issues(
        self,
        issue: str = "INVALID_EMAIL",
        search: str | None = None,
        sort_by: str = "PersonID",
        sort_order: str = "desc",
        severity: str | None = None,
        limit: int = 25,
        offset: int = 0,
    ) -> ContactQualityIssuesResponse:
        norm_issue = issue.upper().strip()
        params: dict[str, Any] = {
            "limit": max(1, min(100, limit)),
            "offset": max(0, offset),
        }

        search_where = ""
        if search and search.strip():
            params["search_pattern"] = f"%{search.strip()}%"
            search_where = (
                " AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern "
                " OR c.TypeValue LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)"
            )

        order_by = _resolve_order_clause(sort_by, sort_order, "p.PersonID DESC")

        cte_sql = """
        WITH ClassifiedContacts AS (
            SELECT 
                c.PersonPhoneID,
                c.PersionID AS PersonID,
                c.LabelTypeID,
                l.LableValue AS LabelName,
                c.TypeValue,
                c.IsVerified,
                c.IsPrimary,
                c.PersonPhoneIsActive,
                CASE 
                    WHEN l.LabelType = 'EMail' OR (c.LabelTypeID IS NULL AND c.TypeValue LIKE '%@%') THEN 'EMAIL'
                    WHEN l.LabelType = 'PhoneNumbers' OR (c.LabelTypeID IS NULL AND c.TypeValue NOT LIKE '%@%' AND c.TypeValue NOT LIKE 'http%' AND c.TypeValue NOT LIKE 'www%') THEN 'PHONE'
                    WHEN l.LabelType = 'URL' OR (c.LabelTypeID IS NULL AND (c.TypeValue LIKE 'http%' OR c.TypeValue LIKE 'www%')) THEN 'URL'
                    ELSE 'OTHER'
                END AS ContactCategory,
                LOWER(LTRIM(RTRIM(c.TypeValue))) AS NormalizedEmail,
                REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(LTRIM(RTRIM(c.TypeValue)), ' ', ''), '-', ''), '+', ''), '(', ''), ')', ''), '.', ''), '/', '') AS NormalizedPhone
            FROM dbo.DLPersonPhoneEmailURLDet c
            JOIN dbo.DLPersonMst p ON p.PersonID = c.PersionID
            LEFT JOIN dbo.DLLabelTypeMst l ON c.LabelTypeID = l.LabelTypeID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0
        )
        """

        # ── 1. Contact Quality Handlers ─────────────────────────────
        if norm_issue == ContactQualityIssueType.MISSING_EMAIL.value:
            search_where_person = (
                " AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)"
                if (search and search.strip())
                else ""
            )
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 
              AND ISNULL(p.PersonIsDeleted, 0) = 0 
              AND NOT {QUALIFYING_EMAIL_EXISTS_SQL}
              {search_where_person};
            """
            items_sql = f"""
            SELECT 
                p.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                NULL AS ContactID,
                'EMAIL' AS ContactType,
                NULL AS LabelName,
                NULL AS CurrentValue,
                'MISSING_EMAIL' AS IssueCode,
                'Person record does not have any registered email address' AS IssueDescription,
                'WARNING' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 
              AND ISNULL(p.PersonIsDeleted, 0) = 0 
              AND NOT {QUALIFYING_EMAIL_EXISTS_SQL}
              {search_where_person}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.MISSING_PHONE.value:
            search_where_person = (
                " AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)"
                if (search and search.strip())
                else ""
            )
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 
              AND ISNULL(p.PersonIsDeleted, 0) = 0 
              AND NOT {QUALIFYING_PHONE_EXISTS_SQL}
              {search_where_person};
            """
            items_sql = f"""
            SELECT 
                p.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                NULL AS ContactID,
                'PHONE' AS ContactType,
                NULL AS LabelName,
                NULL AS CurrentValue,
                'MISSING_PHONE' AS IssueCode,
                'Person record does not have any registered phone number' AS IssueDescription,
                'WARNING' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 
              AND ISNULL(p.PersonIsDeleted, 0) = 0 
              AND NOT {QUALIFYING_PHONE_EXISTS_SQL}
              {search_where_person}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.INVALID_EMAIL.value:
            base_where = """
            c.ContactCategory = 'EMAIL' AND (
                c.TypeValue IS NULL OR LTRIM(RTRIM(c.TypeValue)) = '' 
                OR c.TypeValue NOT LIKE '%_@__%.__%'
                OR c.TypeValue LIKE '% %'
                OR c.TypeValue LIKE '%@%@%'
                OR c.TypeValue LIKE '%.@%'
                OR c.TypeValue LIKE '%@.%'
            )
            """
            count_sql = f"{cte_sql} SELECT COUNT_BIG(1) AS total FROM ClassifiedContacts c JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID WHERE {base_where} {search_where};"
            items_sql = f"""
            {cte_sql}
            SELECT 
                c.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                c.PersonPhoneID AS ContactID,
                'EMAIL' AS ContactType,
                c.LabelName AS LabelName,
                c.TypeValue AS CurrentValue,
                'INVALID_EMAIL' AS IssueCode,
                'Malformed email format or invalid characters' AS IssueDescription,
                'CRITICAL' AS Severity,
                c.IsVerified AS IsVerified,
                c.IsPrimary AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM ClassifiedContacts c
            JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID
            WHERE {base_where} {search_where}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.INVALID_PHONE.value:
            base_where = """
            c.ContactCategory = 'PHONE' AND (
                c.TypeValue IS NULL OR LTRIM(RTRIM(c.TypeValue)) = ''
                OR (
                    CASE 
                        WHEN LOWER(ISNULL(c.LabelName, '')) LIKE '%ext%' THEN
                            CASE WHEN LEN(c.NormalizedPhone) = 4 AND c.NormalizedPhone NOT LIKE '%[^0-9]%' THEN 0 ELSE 1 END
                        ELSE
                            CASE WHEN LEN(c.NormalizedPhone) < 7 OR LEN(c.NormalizedPhone) > 15 OR c.NormalizedPhone LIKE '%[^0-9]%' OR c.NormalizedPhone IN ('0000000000', '1234567890', '9999999999', '1111111111') THEN 1 ELSE 0 END
                    END = 1
                )
            )
            """
            count_sql = f"{cte_sql} SELECT COUNT_BIG(1) AS total FROM ClassifiedContacts c JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID WHERE {base_where} {search_where};"
            items_sql = f"""
            {cte_sql}
            SELECT 
                c.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                c.PersonPhoneID AS ContactID,
                'PHONE' AS ContactType,
                c.LabelName AS LabelName,
                c.TypeValue AS CurrentValue,
                'INVALID_PHONE' AS IssueCode,
                CASE 
                    WHEN LOWER(ISNULL(c.LabelName, '')) LIKE '%ext%' THEN 'Extension number must be exactly 4 numeric digits'
                    ELSE 'Phone number has invalid length (7-15 digits), alpha characters, or dummy value'
                END AS IssueDescription,
                'CRITICAL' AS Severity,
                c.IsVerified AS IsVerified,
                c.IsPrimary AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM ClassifiedContacts c
            JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID
            WHERE {base_where} {search_where}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.INVALID_URL.value:
            base_where = """
            c.ContactCategory = 'URL' AND (
                c.TypeValue IS NULL OR LTRIM(RTRIM(c.TypeValue)) = ''
                OR (c.TypeValue NOT LIKE 'http://%' AND c.TypeValue NOT LIKE 'https://%' AND c.TypeValue NOT LIKE 'www.%')
            )
            """
            count_sql = f"{cte_sql} SELECT COUNT_BIG(1) AS total FROM ClassifiedContacts c JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID WHERE {base_where} {search_where};"
            items_sql = f"""
            {cte_sql}
            SELECT 
                c.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                c.PersonPhoneID AS ContactID,
                'URL' AS ContactType,
                c.LabelName AS LabelName,
                c.TypeValue AS CurrentValue,
                'INVALID_URL' AS IssueCode,
                'URL does not start with valid scheme (http://, https://, or www.)' AS IssueDescription,
                'WARNING' AS Severity,
                c.IsVerified AS IsVerified,
                c.IsPrimary AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM ClassifiedContacts c
            JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID
            WHERE {base_where} {search_where}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.UNVERIFIED_CONTACT.value:
            base_where = "c.IsVerified = 0 OR c.IsVerified IS NULL"
            count_sql = f"{cte_sql} SELECT COUNT_BIG(1) AS total FROM ClassifiedContacts c JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID WHERE {base_where} {search_where};"
            items_sql = f"""
            {cte_sql}
            SELECT 
                c.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                c.PersonPhoneID AS ContactID,
                c.ContactCategory AS ContactType,
                c.LabelName AS LabelName,
                c.TypeValue AS CurrentValue,
                'UNVERIFIED_CONTACT' AS IssueCode,
                'Contact channel has not been verified' AS IssueDescription,
                'INFO' AS Severity,
                c.IsVerified AS IsVerified,
                c.IsPrimary AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM ClassifiedContacts c
            JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID
            WHERE {base_where} {search_where}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue in (ContactQualityIssueType.DUPLICATE_EMAIL_CROSS.value, "DUPLICATE_EMAIL"):
            count_sql = f"""
            {cte_sql}
            SELECT COUNT_BIG(1) AS total 
            FROM ClassifiedContacts c
            JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID
            WHERE c.ContactCategory = 'EMAIL' 
              AND c.NormalizedEmail IN (
                  SELECT NormalizedEmail 
                  FROM ClassifiedContacts 
                  WHERE ContactCategory = 'EMAIL' AND NormalizedEmail <> '' AND NormalizedEmail IS NOT NULL
                  GROUP BY NormalizedEmail 
                  HAVING COUNT(DISTINCT PersonID) > 1
              )
              {search_where};
            """
            items_sql = f"""
            {cte_sql}
            SELECT 
                c.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                c.PersonPhoneID AS ContactID,
                'EMAIL' AS ContactType,
                c.LabelName AS LabelName,
                c.TypeValue AS CurrentValue,
                'DUPLICATE_EMAIL_CROSS' AS IssueCode,
                'Email is shared across multiple distinct Person accounts' AS IssueDescription,
                'WARNING' AS Severity,
                c.IsVerified AS IsVerified,
                c.IsPrimary AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM ClassifiedContacts c
            JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID
            WHERE c.ContactCategory = 'EMAIL' 
              AND c.NormalizedEmail IN (
                  SELECT NormalizedEmail 
                  FROM ClassifiedContacts 
                  WHERE ContactCategory = 'EMAIL' AND NormalizedEmail <> '' AND NormalizedEmail IS NOT NULL
                  GROUP BY NormalizedEmail 
                  HAVING COUNT(DISTINCT PersonID) > 1
              )
              {search_where}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.DUPLICATE_EMAIL_SAME.value:
            count_sql = f"""
            {cte_sql}
            SELECT COUNT_BIG(1) AS total 
            FROM ClassifiedContacts c
            JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID
            WHERE c.ContactCategory = 'EMAIL' 
              AND EXISTS (
                  SELECT 1 FROM ClassifiedContacts c2 
                  WHERE c2.PersonID = c.PersonID 
                    AND c2.NormalizedEmail = c.NormalizedEmail 
                    AND c2.PersonPhoneID <> c.PersonPhoneID
              )
              {search_where};
            """
            items_sql = f"""
            {cte_sql}
            SELECT 
                c.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                c.PersonPhoneID AS ContactID,
                'EMAIL' AS ContactType,
                c.LabelName AS LabelName,
                c.TypeValue AS CurrentValue,
                'DUPLICATE_EMAIL_SAME' AS IssueCode,
                'Duplicate email entered multiple times for this Person' AS IssueDescription,
                'WARNING' AS Severity,
                c.IsVerified AS IsVerified,
                c.IsPrimary AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM ClassifiedContacts c
            JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID
            WHERE c.ContactCategory = 'EMAIL' 
              AND EXISTS (
                  SELECT 1 FROM ClassifiedContacts c2 
                  WHERE c2.PersonID = c.PersonID 
                    AND c2.NormalizedEmail = c.NormalizedEmail 
                    AND c2.PersonPhoneID <> c.PersonPhoneID
              )
              {search_where}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue in (ContactQualityIssueType.DUPLICATE_PHONE_CROSS.value, "DUPLICATE_PHONE"):
            count_sql = f"""
            {cte_sql}
            SELECT COUNT_BIG(1) AS total 
            FROM ClassifiedContacts c
            JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID
            WHERE c.ContactCategory = 'PHONE' AND LEN(c.NormalizedPhone) >= 7
              AND c.NormalizedPhone IN (
                  SELECT NormalizedPhone 
                  FROM ClassifiedContacts 
                  WHERE ContactCategory = 'PHONE' AND NormalizedPhone <> '' AND NormalizedPhone IS NOT NULL AND LEN(NormalizedPhone) >= 7
                  GROUP BY NormalizedPhone 
                  HAVING COUNT(DISTINCT PersonID) > 1
              )
              {search_where};
            """
            items_sql = f"""
            {cte_sql}
            SELECT 
                c.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                c.PersonPhoneID AS ContactID,
                'PHONE' AS ContactType,
                c.LabelName AS LabelName,
                c.TypeValue AS CurrentValue,
                'DUPLICATE_PHONE_CROSS' AS IssueCode,
                'Phone number is shared across multiple distinct Person accounts' AS IssueDescription,
                'WARNING' AS Severity,
                c.IsVerified AS IsVerified,
                c.IsPrimary AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM ClassifiedContacts c
            JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID
            WHERE c.ContactCategory = 'PHONE' AND LEN(c.NormalizedPhone) >= 7
              AND c.NormalizedPhone IN (
                  SELECT NormalizedPhone 
                  FROM ClassifiedContacts 
                  WHERE ContactCategory = 'PHONE' AND NormalizedPhone <> '' AND NormalizedPhone IS NOT NULL AND LEN(NormalizedPhone) >= 7
                  GROUP BY NormalizedPhone 
                  HAVING COUNT(DISTINCT PersonID) > 1
              )
              {search_where}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.DUPLICATE_PHONE_SAME.value:
            count_sql = f"""
            {cte_sql}
            SELECT COUNT_BIG(1) AS total 
            FROM ClassifiedContacts c
            JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID
            WHERE c.ContactCategory = 'PHONE' AND LEN(c.NormalizedPhone) >= 7
              AND EXISTS (
                  SELECT 1 FROM ClassifiedContacts c2 
                  WHERE c2.PersonID = c.PersonID 
                    AND c2.NormalizedPhone = c.NormalizedPhone 
                    AND c2.PersonPhoneID <> c.PersonPhoneID
              )
              {search_where};
            """
            items_sql = f"""
            {cte_sql}
            SELECT 
                c.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                c.PersonPhoneID AS ContactID,
                'PHONE' AS ContactType,
                c.LabelName AS LabelName,
                c.TypeValue AS CurrentValue,
                'DUPLICATE_PHONE_SAME' AS IssueCode,
                'Duplicate phone entered multiple times for this Person' AS IssueDescription,
                'WARNING' AS Severity,
                c.IsVerified AS IsVerified,
                c.IsPrimary AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM ClassifiedContacts c
            JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID
            WHERE c.ContactCategory = 'PHONE' AND LEN(c.NormalizedPhone) >= 7
              AND EXISTS (
                  SELECT 1 FROM ClassifiedContacts c2 
                  WHERE c2.PersonID = c.PersonID 
                    AND c2.NormalizedPhone = c.NormalizedPhone 
                    AND c2.PersonPhoneID <> c.PersonPhoneID
              )
              {search_where}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.MULTIPLE_PRIMARY.value:
            count_sql = f"""
            {cte_sql}
            SELECT COUNT_BIG(1) AS total
            FROM ClassifiedContacts c
            JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID
            WHERE c.IsPrimary = 1
              AND c.PersonID IN (
                  SELECT PersonID
                  FROM ClassifiedContacts
                  WHERE IsPrimary = 1
                  GROUP BY PersonID
                  HAVING COUNT_BIG(1) > 1
              )
              {search_where};
            """
            items_sql = f"""
            {cte_sql}
            SELECT 
                c.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                c.PersonPhoneID AS ContactID,
                c.ContactCategory AS ContactType,
                c.LabelName AS LabelName,
                c.TypeValue AS CurrentValue,
                'MULTIPLE_PRIMARY' AS IssueCode,
                'Person record has multiple contacts flagged as Primary' AS IssueDescription,
                'CRITICAL' AS Severity,
                c.IsVerified AS IsVerified,
                c.IsPrimary AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM ClassifiedContacts c
            JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID
            WHERE c.IsPrimary = 1
              AND c.PersonID IN (
                  SELECT PersonID
                  FROM ClassifiedContacts
                  WHERE IsPrimary = 1
                  GROUP BY PersonID
                  HAVING COUNT_BIG(1) > 1
              )
              {search_where}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.PRIMARY_INACTIVE.value:
            count_sql = f"""
            {cte_sql}
            SELECT COUNT_BIG(1) AS total
            FROM ClassifiedContacts c
            JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID
            WHERE c.IsPrimary = 1 AND (c.PersonPhoneIsActive = 0 OR c.PersonPhoneIsActive IS NULL)
              {search_where};
            """
            items_sql = f"""
            {cte_sql}
            SELECT 
                c.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                c.PersonPhoneID AS ContactID,
                c.ContactCategory AS ContactType,
                c.LabelName AS LabelName,
                c.TypeValue AS CurrentValue,
                'PRIMARY_INACTIVE' AS IssueCode,
                'Primary contact is marked as inactive or disabled' AS IssueDescription,
                'CRITICAL' AS Severity,
                c.IsVerified AS IsVerified,
                c.IsPrimary AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM ClassifiedContacts c
            JOIN dbo.DLPersonMst p ON c.PersonID = p.PersonID
            WHERE c.IsPrimary = 1 AND (c.PersonPhoneIsActive = 0 OR c.PersonPhoneIsActive IS NULL)
              {search_where}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        # ── 2. Address Quality Handlers ─────────────────────────────
        elif norm_issue == ContactQualityIssueType.MISSING_POSTAL_CODE.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonAddressDet a
            JOIN dbo.DLPersonMst p ON a.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (a.PostalCode IS NULL OR LTRIM(RTRIM(a.PostalCode)) = '')
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR a.CityName LIKE :search_pattern OR a.Street LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                a.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                a.PersonAddID AS ContactID,
                'ADDRESS' AS ContactType,
                ISNULL(a.CityName, 'Address') AS LabelName,
                ISNULL(a.Street, a.CityName) AS CurrentValue,
                'MISSING_POSTAL_CODE' AS IssueCode,
                'Address record has no postal / PIN code provided' AS IssueDescription,
                'WARNING' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonAddressDet a
            JOIN dbo.DLPersonMst p ON a.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (a.PostalCode IS NULL OR LTRIM(RTRIM(a.PostalCode)) = '')
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR a.CityName LIKE :search_pattern OR a.Street LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.INVALID_PIN_CODE_FORMAT.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonAddressDet a
            JOIN dbo.DLPersonMst p ON a.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND a.PostalCode IS NOT NULL AND LTRIM(RTRIM(a.PostalCode)) <> '' 
              AND (LEN(LTRIM(RTRIM(a.PostalCode))) NOT IN (6, 5) OR LTRIM(RTRIM(a.PostalCode)) LIKE '%[^0-9]%')
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR a.CityName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                a.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                a.PersonAddID AS ContactID,
                'ADDRESS' AS ContactType,
                ISNULL(a.CityName, 'Postal Code') AS LabelName,
                a.PostalCode AS CurrentValue,
                'INVALID_PIN_CODE_FORMAT' AS IssueCode,
                'Postal code contains non-numeric characters or invalid length' AS IssueDescription,
                'CRITICAL' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonAddressDet a
            JOIN dbo.DLPersonMst p ON a.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND a.PostalCode IS NOT NULL AND LTRIM(RTRIM(a.PostalCode)) <> '' 
              AND (LEN(LTRIM(RTRIM(a.PostalCode))) NOT IN (6, 5) OR LTRIM(RTRIM(a.PostalCode)) LIKE '%[^0-9]%')
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR a.CityName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.STREET_WITHOUT_CITY.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonAddressDet a
            JOIN dbo.DLPersonMst p ON a.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (a.CityName IS NULL OR LTRIM(RTRIM(a.CityName)) = '') AND a.Street IS NOT NULL AND LTRIM(RTRIM(a.Street)) <> ''
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR a.Street LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                a.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                a.PersonAddID AS ContactID,
                'ADDRESS' AS ContactType,
                'Street Address' AS LabelName,
                a.Street AS CurrentValue,
                'STREET_WITHOUT_CITY' AS IssueCode,
                'Street address is present but city name is missing or blank' AS IssueDescription,
                'WARNING' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonAddressDet a
            JOIN dbo.DLPersonMst p ON a.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (a.CityName IS NULL OR LTRIM(RTRIM(a.CityName)) = '') AND a.Street IS NOT NULL AND LTRIM(RTRIM(a.Street)) <> ''
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR a.Street LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.CITY_WITHOUT_STATE.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonAddressDet a
            JOIN dbo.DLPersonMst p ON a.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND a.CityName IS NOT NULL AND LTRIM(RTRIM(a.CityName)) <> '' AND (a.StateName IS NULL OR LTRIM(RTRIM(a.StateName)) = '')
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR a.CityName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                a.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                a.PersonAddID AS ContactID,
                'ADDRESS' AS ContactType,
                'City' AS LabelName,
                a.CityName AS CurrentValue,
                'CITY_WITHOUT_STATE' AS IssueCode,
                'Address has city name but missing state region' AS IssueDescription,
                'WARNING' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonAddressDet a
            JOIN dbo.DLPersonMst p ON a.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND a.CityName IS NOT NULL AND LTRIM(RTRIM(a.CityName)) <> '' AND (a.StateName IS NULL OR LTRIM(RTRIM(a.StateName)) = '')
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR a.CityName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.MISSING_GEOCODES.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonAddressDet a
            JOIN dbo.DLPersonMst p ON a.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (a.Latitude IS NULL OR a.Longitude IS NULL OR (a.Latitude = 0 AND a.Longitude = 0))
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR a.CityName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                a.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                a.PersonAddID AS ContactID,
                'ADDRESS' AS ContactType,
                ISNULL(a.CityName, 'Address') AS LabelName,
                ISNULL(a.GoogleFormattedAddress, a.Street) AS CurrentValue,
                'MISSING_GEOCODES' AS IssueCode,
                'Address has not been geocoded with valid Latitude and Longitude' AS IssueDescription,
                'INFO' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonAddressDet a
            JOIN dbo.DLPersonMst p ON a.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (a.Latitude IS NULL OR a.Longitude IS NULL OR (a.Latitude = 0 AND a.Longitude = 0))
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR a.CityName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.DUPLICATE_ADDRESSES_SAME_PERSON.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonAddressDet a
            JOIN dbo.DLPersonMst p ON a.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND a.Street IS NOT NULL AND LTRIM(RTRIM(a.Street)) <> ''
              AND EXISTS (
                  SELECT 1 FROM dbo.DLPersonAddressDet a2
                  WHERE a2.PersonID = a.PersonID 
                    AND LOWER(LTRIM(RTRIM(a2.Street))) = LOWER(LTRIM(RTRIM(a.Street)))
                    AND a2.PersonAddID <> a.PersonAddID
              )
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR a.Street LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                a.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                a.PersonAddID AS ContactID,
                'ADDRESS' AS ContactType,
                'Duplicate Address' AS LabelName,
                a.Street + ', ' + ISNULL(a.CityName, '') AS CurrentValue,
                'DUPLICATE_ADDRESSES_SAME_PERSON' AS IssueCode,
                'Identical address entered multiple times for this Person' AS IssueDescription,
                'WARNING' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonAddressDet a
            JOIN dbo.DLPersonMst p ON a.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND a.Street IS NOT NULL AND LTRIM(RTRIM(a.Street)) <> ''
              AND EXISTS (
                  SELECT 1 FROM dbo.DLPersonAddressDet a2
                  WHERE a2.PersonID = a.PersonID 
                    AND LOWER(LTRIM(RTRIM(a2.Street))) = LOWER(LTRIM(RTRIM(a.Street)))
                    AND a2.PersonAddID <> a.PersonAddID
              )
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR a.Street LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        # ── 3. Profile & Chronological Integrity Handlers ───────────
        elif norm_issue == ContactQualityIssueType.ANNIVERSARY_BEFORE_BIRTH.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND p.PersonAnneversaryDate < p.PersonBirthDate
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                p.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                NULL AS ContactID,
                'PROFILE' AS ContactType,
                'Anniversary Date' AS LabelName,
                CONVERT(varchar, p.PersonAnneversaryDate, 23) + ' < ' + CONVERT(varchar, p.PersonBirthDate, 23) AS CurrentValue,
                'ANNIVERSARY_BEFORE_BIRTH' AS IssueCode,
                'Anniversary date is earlier than birth date (chronological error)' AS IssueDescription,
                'CRITICAL' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND p.PersonAnneversaryDate < p.PersonBirthDate
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.INVALID_BIRTH_DATE.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (p.PersonBirthDate > GETDATE() OR p.PersonBirthDate < '1900-01-01')
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                p.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                NULL AS ContactID,
                'PROFILE' AS ContactType,
                'Birth Date' AS LabelName,
                CONVERT(varchar, p.PersonBirthDate, 23) AS CurrentValue,
                'INVALID_BIRTH_DATE' AS IssueCode,
                'Birth date is in the future or before 1900' AS IssueDescription,
                'CRITICAL' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (p.PersonBirthDate > GETDATE() OR p.PersonBirthDate < '1900-01-01')
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.BIRTH_DATE_DEFAULT_OR_ANCIENT.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND p.PersonBirthDate IS NOT NULL AND (DATEDIFF(year, p.PersonBirthDate, GETDATE()) < 0 OR DATEDIFF(year, p.PersonBirthDate, GETDATE()) > 100)
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                p.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                NULL AS ContactID,
                'PROFILE' AS ContactType,
                'Birth Date' AS LabelName,
                CONVERT(varchar, p.PersonBirthDate, 23) AS CurrentValue,
                'BIRTH_DATE_DEFAULT_OR_ANCIENT' AS IssueCode,
                'Birth date is set to dummy 1900-01-01 or age > 100 years' AS IssueDescription,
                'WARNING' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND p.PersonBirthDate IS NOT NULL AND (DATEDIFF(year, p.PersonBirthDate, GETDATE()) < 0 OR DATEDIFF(year, p.PersonBirthDate, GETDATE()) > 100)
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.SUSPICIOUS_DUMMY_NAMES.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (
                LOWER(p.PersonFirstName) IN ('test', 'admin', 'dummy', 'asdf', 'xyz', 'na', 'n/a', 'none', 'null', 'temp') 
                OR LOWER(p.PersonLastName) IN ('test', 'admin', 'dummy', 'asdf', 'xyz', 'na', 'n/a', 'none', 'null', 'temp')
            )
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                p.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                NULL AS ContactID,
                'PROFILE' AS ContactType,
                'Suspect Name' AS LabelName,
                ISNULL(p.PersonFirstName, '') + ' ' + ISNULL(p.PersonLastName, '') AS CurrentValue,
                'SUSPICIOUS_DUMMY_NAMES' AS IssueCode,
                'Person record has placeholder/test name (test, admin, dummy, etc.)' AS IssueDescription,
                'WARNING' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (
                LOWER(p.PersonFirstName) IN ('test', 'admin', 'dummy', 'asdf', 'xyz', 'na', 'n/a', 'none', 'null', 'temp') 
                OR LOWER(p.PersonLastName) IN ('test', 'admin', 'dummy', 'asdf', 'xyz', 'na', 'n/a', 'none', 'null', 'temp')
            )
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.MISSING_LAST_NAME.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (p.PersonLastName IS NULL OR LTRIM(RTRIM(p.PersonLastName)) = '') AND p.PersonFirstName IS NOT NULL AND LTRIM(RTRIM(p.PersonFirstName)) <> ''
            {"AND (p.PersonFirstName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                p.PersonID,
                p.PersonFirstName AS PersonName,
                NULL AS ContactID,
                'PROFILE' AS ContactType,
                'First Name Only' AS LabelName,
                p.PersonFirstName AS CurrentValue,
                'MISSING_LAST_NAME' AS IssueCode,
                'Person record has first name but missing last name / surname' AS IssueDescription,
                'WARNING' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (p.PersonLastName IS NULL OR LTRIM(RTRIM(p.PersonLastName)) = '') AND p.PersonFirstName IS NOT NULL AND LTRIM(RTRIM(p.PersonFirstName)) <> ''
            {"AND (p.PersonFirstName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        # ── 4. Employment & Lifecycle Consistency Handlers ──────────
        elif norm_issue == ContactQualityIssueType.ACTIVE_EMP_MISSING_TITLE.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonMst p
            WHERE p.EmpID IS NOT NULL AND p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (p.PersonTitle IS NULL OR LTRIM(RTRIM(p.PersonTitle)) = '')
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                p.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                NULL AS ContactID,
                'EMPLOYMENT' AS ContactType,
                'EmpID: ' + CAST(p.EmpID AS varchar) AS LabelName,
                ISNULL(p.PersonDepartment, 'Active Employee') AS CurrentValue,
                'ACTIVE_EMP_MISSING_TITLE' AS IssueCode,
                'Active employee record has no job title designation defined' AS IssueDescription,
                'WARNING' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonMst p
            WHERE p.EmpID IS NOT NULL AND p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (p.PersonTitle IS NULL OR LTRIM(RTRIM(p.PersonTitle)) = '')
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.INACTIVE_WITH_ACTIVE_EMPID.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonMst p
            WHERE (p.PersonIsActive = 0 OR p.PersonIsActive IS NULL) AND p.EmpID IS NOT NULL
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                p.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                NULL AS ContactID,
                'EMPLOYMENT' AS ContactType,
                'EmpID: ' + CAST(p.EmpID AS varchar) AS LabelName,
                'PersonIsActive = 0' AS CurrentValue,
                'INACTIVE_WITH_ACTIVE_EMPID' AS IssueCode,
                'Inactive person record still has an employee ID assigned' AS IssueDescription,
                'INFO' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonMst p
            WHERE (p.PersonIsActive = 0 OR p.PersonIsActive IS NULL) AND p.EmpID IS NOT NULL
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.STATUS_ACTIVE_AND_DELETED.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND p.PersonIsDeleted = 1
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                p.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                NULL AS ContactID,
                'LIFECYCLE' AS ContactType,
                'Status Conflict' AS LabelName,
                'Active=1 AND Deleted=1' AS CurrentValue,
                'STATUS_ACTIVE_AND_DELETED' AS IssueCode,
                'Record has conflicting flags: both active and deleted simultaneously' AS IssueDescription,
                'CRITICAL' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND p.PersonIsDeleted = 1
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.STALE_TEMP_PERSONS.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND p.PersonIsTemp = 1 AND p.PersonEntDt < DATEADD(day, -90, GETDATE())
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                p.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                NULL AS ContactID,
                'LIFECYCLE' AS ContactType,
                'Temp Person' AS LabelName,
                'Created: ' + CONVERT(varchar, p.PersonEntDt, 23) AS CurrentValue,
                'STALE_TEMP_PERSONS' AS IssueCode,
                'Temporary person record was created over 90 days ago and not finalized' AS IssueDescription,
                'WARNING' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND p.PersonIsTemp = 1 AND p.PersonEntDt < DATEADD(day, -90, GETDATE())
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        # ── 5. Governance & Blacklist Handlers ───────────────────────
        elif norm_issue == ContactQualityIssueType.BLACKLIST_UNAPPROVED.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND p.PersonIsBlackList = 1 AND (p.PersonBlackListHODApprove = 0 OR p.PersonBlackListHODApprove IS NULL)
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                p.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                NULL AS ContactID,
                'GOVERNANCE' AS ContactType,
                'Blacklist Status' AS LabelName,
                ISNULL(p.PersonBlackListType, 'Blacklisted') AS CurrentValue,
                'BLACKLIST_UNAPPROVED' AS IssueCode,
                'Person is flagged as blacklisted without HOD approval' AS IssueDescription,
                'CRITICAL' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND p.PersonIsBlackList = 1 AND (p.PersonBlackListHODApprove = 0 OR p.PersonBlackListHODApprove IS NULL)
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.BLACKLIST_MISSING_DETAILS.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND p.PersonIsBlackList = 1 AND (p.PersonBlackListDate IS NULL OR p.PersonBlackListType IS NULL)
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                p.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                NULL AS ContactID,
                'GOVERNANCE' AS ContactType,
                'Blacklist Reason' AS LabelName,
                'Missing Date/Type' AS CurrentValue,
                'BLACKLIST_MISSING_DETAILS' AS IssueCode,
                'Blacklist record is missing either the date or reason category' AS IssueDescription,
                'CRITICAL' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND p.PersonIsBlackList = 1 AND (p.PersonBlackListDate IS NULL OR p.PersonBlackListType IS NULL)
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        # ── 6. Entity Linkages & Child Records Handlers ─────────────
        elif norm_issue == ContactQualityIssueType.ORPHAN_COMPANY_LINK.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonCompanyLinkDet l
            JOIN dbo.DLPersonMst p ON l.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND NOT EXISTS (SELECT 1 FROM dbo.DLCompanyMst c WHERE c.DLCompID = l.DLCompID)
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                l.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                l.PersonLinkID AS ContactID,
                'COMPANY' AS ContactType,
                'Company Link' AS LabelName,
                'DLCompID: ' + CAST(ISNULL(l.DLCompID, 0) AS varchar) AS CurrentValue,
                'ORPHAN_COMPANY_LINK' AS IssueCode,
                'Company link references a missing or non-existent Company ID' AS IssueDescription,
                'CRITICAL' AS Severity,
                NULL AS IsVerified,
                l.IsPrimary AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonCompanyLinkDet l
            JOIN dbo.DLPersonMst p ON l.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND NOT EXISTS (SELECT 1 FROM dbo.DLCompanyMst c WHERE c.DLCompID = l.DLCompID)
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.DUPLICATE_COMPANY_LINKS.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonCompanyLinkDet l
            JOIN dbo.DLPersonMst p ON l.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND l.DLCompID IS NOT NULL
              AND EXISTS (
                  SELECT 1 FROM dbo.DLPersonCompanyLinkDet l2
                  WHERE l2.PersonID = l.PersonID 
                    AND l2.DLCompID = l.DLCompID
                    AND l2.PersonLinkID <> l.PersonLinkID
              )
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                l.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                l.PersonLinkID AS ContactID,
                'COMPANY' AS ContactType,
                ISNULL(c.DLCompName, 'Company') AS LabelName,
                'DLCompID: ' + CAST(l.DLCompID AS varchar) AS CurrentValue,
                'DUPLICATE_COMPANY_LINKS' AS IssueCode,
                'Identical company linked multiple times for this Person' AS IssueDescription,
                'WARNING' AS Severity,
                NULL AS IsVerified,
                l.IsPrimary AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonCompanyLinkDet l
            JOIN dbo.DLPersonMst p ON l.PersonID = p.PersonID
            LEFT JOIN dbo.DLCompanyMst c ON l.DLCompID = c.DLCompID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND l.DLCompID IS NOT NULL
              AND EXISTS (
                  SELECT 1 FROM dbo.DLPersonCompanyLinkDet l2
                  WHERE l2.PersonID = l.PersonID 
                    AND l2.DLCompID = l.DLCompID
                    AND l2.PersonLinkID <> l.PersonLinkID
              )
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.COMPANY_MISSING_ROLE.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonCompanyLinkDet l
            JOIN dbo.DLPersonMst p ON l.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (l.CompPersonRoleID IS NULL OR l.CompPersonRoleID = 0)
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                l.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                l.PersonLinkID AS ContactID,
                'COMPANY' AS ContactType,
                ISNULL(c.DLCompName, 'Company Link') AS LabelName,
                'CompPersonRoleID = NULL' AS CurrentValue,
                'COMPANY_MISSING_ROLE' AS IssueCode,
                'Company affiliation link is missing designation role' AS IssueDescription,
                'WARNING' AS Severity,
                NULL AS IsVerified,
                l.IsPrimary AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonCompanyLinkDet l
            JOIN dbo.DLPersonMst p ON l.PersonID = p.PersonID
            LEFT JOIN dbo.DLCompanyMst c ON l.DLCompID = c.DLCompID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND (l.CompPersonRoleID IS NULL OR l.CompPersonRoleID = 0)
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.EXTRA_FIELD_ORPHAN_ID.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonExtraFieldValueDet e
            JOIN dbo.DLPersonMst p ON e.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND NOT EXISTS (SELECT 1 FROM dbo.DLExtraFieldDet m WHERE m.ExtraFieldID = e.ExtraFieldID)
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                e.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                e.PersonExtraFieldValueID AS ContactID,
                'CUSTOM_FIELD' AS ContactType,
                'ExtraFieldID: ' + CAST(ISNULL(e.ExtraFieldID, 0) AS varchar) AS LabelName,
                e.PersonExtraFieldValue AS CurrentValue,
                'EXTRA_FIELD_ORPHAN_ID' AS IssueCode,
                'Custom field value references a missing or non-existent ExtraField definition' AS IssueDescription,
                'CRITICAL' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonExtraFieldValueDet e
            JOIN dbo.DLPersonMst p ON e.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND NOT EXISTS (SELECT 1 FROM dbo.DLExtraFieldDet m WHERE m.ExtraFieldID = e.ExtraFieldID)
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.DUPLICATE_EXTRA_FIELDS.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonExtraFieldValueDet e
            JOIN dbo.DLPersonMst p ON e.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND EXISTS (
                SELECT 1 FROM dbo.DLPersonExtraFieldValueDet e2
                WHERE e2.PersonID = e.PersonID 
                  AND e2.ExtraFieldID = e.ExtraFieldID
                  AND e2.PersonExtraFieldValueID <> e.PersonExtraFieldValueID
            )
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                e.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                e.PersonExtraFieldValueID AS ContactID,
                'CUSTOM_FIELD' AS ContactType,
                'ExtraFieldID: ' + CAST(e.ExtraFieldID AS varchar) AS LabelName,
                e.PersonExtraFieldValue AS CurrentValue,
                'DUPLICATE_EXTRA_FIELDS' AS IssueCode,
                'Duplicate custom attribute value entered multiple times for this Person' AS IssueDescription,
                'WARNING' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonExtraFieldValueDet e
            JOIN dbo.DLPersonMst p ON e.PersonID = p.PersonID
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND EXISTS (
                SELECT 1 FROM dbo.DLPersonExtraFieldValueDet e2
                WHERE e2.PersonID = e.PersonID 
                  AND e2.ExtraFieldID = e.ExtraFieldID
                  AND e2.PersonExtraFieldValueID <> e.PersonExtraFieldValueID
            )
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        # ── 7. Audit Trail & Sync Integration Handlers ──────────────
        elif norm_issue == ContactQualityIssueType.DELETED_MISSING_TIMESTAMP.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsDeleted = 1 AND p.PersonDelDt IS NULL
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                p.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                NULL AS ContactID,
                'AUDIT' AS ContactType,
                'Deletion Audit' AS LabelName,
                'DelDt IS NULL' AS CurrentValue,
                'DELETED_MISSING_TIMESTAMP' AS IssueCode,
                'Record is marked as deleted but has no deletion timestamp' AS IssueDescription,
                'WARNING' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsDeleted = 1 AND p.PersonDelDt IS NULL
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.AUDIT_DEL_BEFORE_ENT.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonMst p
            WHERE p.PersonDelDt IS NOT NULL AND p.PersonEntDt IS NOT NULL AND p.PersonDelDt < p.PersonEntDt
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                p.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                NULL AS ContactID,
                'AUDIT' AS ContactType,
                'Chronological Error' AS LabelName,
                'Del: ' + CONVERT(varchar, p.PersonDelDt, 23) + ' < Ent: ' + CONVERT(varchar, p.PersonEntDt, 23) AS CurrentValue,
                'AUDIT_DEL_BEFORE_ENT' AS IssueCode,
                'Deletion date is earlier than creation date (audit trail corruption)' AS IssueDescription,
                'CRITICAL' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonMst p
            WHERE p.PersonDelDt IS NOT NULL AND p.PersonEntDt IS NOT NULL AND p.PersonDelDt < p.PersonEntDt
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        elif norm_issue == ContactQualityIssueType.SYNC_ZIMBRA_MISSING_ID.value:
            count_sql = f"""
            SELECT COUNT_BIG(1) AS total
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND p.IsContactSync = 1 AND p.ZimbraContactID IS NULL
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""};
            """
            items_sql = f"""
            SELECT 
                p.PersonID,
                {PERSON_NAME_SQL} AS PersonName,
                NULL AS ContactID,
                'SYNC' AS ContactType,
                'Zimbra Sync' AS LabelName,
                'ZimbraContactID = NULL' AS CurrentValue,
                'SYNC_ZIMBRA_MISSING_ID' AS IssueCode,
                'Sync is enabled on record but Zimbra Contact ID is missing' AS IssueDescription,
                'WARNING' AS Severity,
                NULL AS IsVerified,
                NULL AS IsPrimary,
                p.PersonIsActive AS IsActive
            FROM dbo.DLPersonMst p
            WHERE p.PersonIsActive = 1 AND ISNULL(p.PersonIsDeleted, 0) = 0 AND p.IsContactSync = 1 AND p.ZimbraContactID IS NULL
            {"AND (p.PersonFirstName LIKE :search_pattern OR p.PersonLastName LIKE :search_pattern OR CAST(p.PersonID AS varchar) LIKE :search_pattern)" if search else ""}
            ORDER BY {order_by}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
            """

        else:
            return ContactQualityIssuesResponse(
                issue=norm_issue,
                total=0,
                limit=limit,
                offset=offset,
                items=[],
            )

        count_rows = execute_readonly_query(count_sql, params)
        total_count = int(count_rows[0]["total"]) if count_rows else 0

        if total_count == 0:
            return ContactQualityIssuesResponse(
                issue=norm_issue,
                total=0,
                limit=limit,
                offset=offset,
                items=[],
            )

        rows = execute_readonly_query(items_sql, params)
        items: list[ContactQualityIssueItem] = []
        for r in rows:
            raw_val = r.get("CurrentValue")
            c_type = r.get("ContactType") or "CONTACT"
            masked = mask_contact_value(c_type, raw_val) if raw_val else "—"

            item = ContactQualityIssueItem(
                PersonID=int(r["PersonID"]),
                PersonName=str(r["PersonName"]),
                ContactID=int(r["ContactID"]) if r.get("ContactID") else None,
                ContactType=c_type,
                LabelName=r.get("LabelName"),
                CurrentValue=raw_val,
                MaskedValue=masked,
                IssueCode=str(r["IssueCode"]),
                IssueDescription=str(r["IssueDescription"]),
                Severity=str(r.get("Severity") or "WARNING"),
                IsVerified=bool(r["IsVerified"]) if r.get("IsVerified") is not None else None,
                IsPrimary=bool(r["IsPrimary"]) if r.get("IsPrimary") is not None else None,
                IsActive=bool(r["IsActive"]) if r.get("IsActive") is not None else None,
            )
            items.append(item)

        return ContactQualityIssuesResponse(
            issue=norm_issue,
            total=total_count,
            limit=limit,
            offset=offset,
            items=items,
        )

    async def export_contact_quality_issues(
        self,
        issue: str = "INVALID_EMAIL",
        format: str = "xlsx",
        search: str | None = None,
        sort_by: str = "PersonID",
        sort_order: str = "desc",
        severity: str | None = None,
    ) -> tuple[bytes, str, str]:
        """
        Exports all matching records for a quality issue as CSV or Excel.
        Returns (content_bytes, media_type, filename).
        """
        norm_issue = issue.upper().strip()
        is_csv = format.lower().strip() == "csv"
        ext = "csv" if is_csv else "xlsx"
        media_type = "text/csv; charset=utf-8" if is_csv else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = f"daylite_{norm_issue.lower()}_{date_str}.{ext}"

        # Fetch all matching rows (up to 50,000 records)
        res = await self.get_contact_quality_issues(
            issue=norm_issue,
            search=search,
            sort_by=sort_by,
            sort_order=sort_order,
            severity=severity,
            limit=50000,
            offset=0,
        )

        headers = [
            "Person ID",
            "Person Name",
            "Issue Code",
            "Severity",
            "Contact Type",
            "Field Label",
            "Offending Value",
            "Issue Description",
            "Is Active",
        ]

        data_rows: list[list[Any]] = []
        for item in res.items:
            data_rows.append([
                item.person_id,
                item.person_name,
                item.issue_code,
                item.severity,
                item.contact_type,
                item.label_name or "",
                item.masked_value or item.current_value or "",
                item.issue_description,
                "Yes" if item.is_active else "No",
            ])

        if is_csv:
            content = generate_csv(headers, data_rows)
        else:
            content = generate_xlsx(norm_issue[:31], headers, data_rows)

        return content, media_type, filename

    async def export_contact_quality_summary(
        self,
        format: str = "xlsx",
    ) -> tuple[bytes, str, str]:
        """
        Exports summary report of all 37 quality rules as CSV or Excel.
        Returns (content_bytes, media_type, filename).
        """
        is_csv = format.lower().strip() == "csv"
        ext = "csv" if is_csv else "xlsx"
        media_type = "text/csv; charset=utf-8" if is_csv else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = f"daylite_quality_summary_{date_str}.{ext}"

        summary = await self.get_contact_quality_summary()

        headers = [
            "Quality Dimension",
            "Rule Name",
            "Issue Code",
            "Severity",
            "Affected Records",
            "Description",
        ]

        rules_map = [
            # 1. Contacts
            ("Contacts & Duplicates", "Missing Email", "MISSING_EMAIL", "WARNING", summary.persons_without_email, "Persons without registered email"),
            ("Contacts & Duplicates", "Missing Phone", "MISSING_PHONE", "WARNING", summary.persons_without_phone, "Persons without registered phone"),
            ("Contacts & Duplicates", "Invalid Email", "INVALID_EMAIL", "CRITICAL", summary.invalid_emails, "Malformed email address syntax"),
            ("Contacts & Duplicates", "Invalid Phone", "INVALID_PHONE", "CRITICAL", summary.invalid_phones, "Invalid phone number length or chars"),
            ("Contacts & Duplicates", "Shared Email", "DUPLICATE_EMAIL_CROSS", "WARNING", summary.duplicate_email_cross_persons, "Identical email shared across persons"),
            ("Contacts & Duplicates", "Shared Phone", "DUPLICATE_PHONE_CROSS", "WARNING", summary.duplicate_phone_cross_persons, "Identical phone shared across persons"),
            ("Contacts & Duplicates", "Duplicate Email (Self)", "DUPLICATE_EMAIL_SAME", "WARNING", summary.duplicate_email_same_person, "Duplicate email under same Person"),
            ("Contacts & Duplicates", "Duplicate Phone (Self)", "DUPLICATE_PHONE_SAME", "WARNING", summary.duplicate_phone_same_person, "Duplicate phone under same Person"),
            ("Contacts & Duplicates", "Unverified Contacts", "UNVERIFIED_CONTACT", "INFO", summary.unverified_contacts, "Contact channels lacking verification"),
            ("Contacts & Duplicates", "Invalid URLs", "INVALID_URL", "WARNING", summary.invalid_urls, "URLs lacking standard web scheme"),
            ("Contacts & Duplicates", "Multiple Primary", "MULTIPLE_PRIMARY", "CRITICAL", summary.persons_multiple_primary, "Multiple primary contacts flagged"),
            ("Contacts & Duplicates", "Primary Inactive", "PRIMARY_INACTIVE", "CRITICAL", summary.primary_contact_inactive, "Primary contact flagged inactive"),

            # 2. Addresses
            ("Address & Locations", "Missing Postal Code", "MISSING_POSTAL_CODE", "WARNING", summary.addr_missing_postal_code, "Address missing PIN / Postal code"),
            ("Address & Locations", "Invalid PIN Format", "INVALID_PIN_CODE_FORMAT", "CRITICAL", summary.addr_invalid_pin_format, "Postal code with non-numeric chars"),
            ("Address & Locations", "Street Without City", "STREET_WITHOUT_CITY", "WARNING", summary.addr_street_without_city, "Street present but city name blank"),
            ("Address & Locations", "City Without State", "CITY_WITHOUT_STATE", "WARNING", summary.addr_city_without_state, "City present but state region blank"),
            ("Address & Locations", "Missing Geocodes", "MISSING_GEOCODES", "INFO", summary.addr_missing_geocodes, "Address lacking Lat/Long coordinates"),
            ("Address & Locations", "Duplicate Address", "DUPLICATE_ADDRESSES_SAME_PERSON", "WARNING", summary.addr_duplicate_same_person, "Identical address repeated for Person"),

            # 3. Profile & Chronology
            ("Profile & Chronology", "Anniversary Before Birth", "ANNIVERSARY_BEFORE_BIRTH", "CRITICAL", summary.person_anniversary_before_birth, "Anniversary earlier than birth date"),
            ("Profile & Chronology", "Invalid Birth Date", "INVALID_BIRTH_DATE", "CRITICAL", summary.person_invalid_birth_date, "Birth date in future or before 1900"),
            ("Profile & Chronology", "Dummy / Ancient DOB", "BIRTH_DATE_DEFAULT_OR_ANCIENT", "WARNING", summary.person_birth_date_ancient, "Dummy 1900-01-01 or age > 100"),
            ("Profile & Chronology", "Suspicious Test Names", "SUSPICIOUS_DUMMY_NAMES", "WARNING", summary.person_suspicious_dummy_names, "Placeholder names (test, dummy, etc.)"),
            ("Profile & Chronology", "Missing Last Name", "MISSING_LAST_NAME", "WARNING", summary.person_missing_lastname_only, "First name present but surname missing"),

            # 4. Governance & Links
            ("Governance & Links", "Active & Deleted Conflict", "STATUS_ACTIVE_AND_DELETED", "CRITICAL", summary.status_active_and_deleted, "Record marked both Active and Deleted"),
            ("Governance & Links", "Employee Missing Title", "ACTIVE_EMP_MISSING_TITLE", "WARNING", summary.active_emp_missing_title, "Active employee missing job title"),
            ("Governance & Links", "Inactive with EmpID", "INACTIVE_WITH_ACTIVE_EMPID", "INFO", summary.inactive_with_empid, "Inactive record retaining employee ID"),
            ("Governance & Links", "Stale Temp Persons", "STALE_TEMP_PERSONS", "WARNING", summary.stale_temp_persons, "Temporary person older than 90 days"),
            ("Governance & Links", "Unapproved Blacklist", "BLACKLIST_UNAPPROVED", "CRITICAL", summary.blacklist_unapproved, "Blacklist active without HOD approval"),
            ("Governance & Links", "Blacklist Missing Details", "BLACKLIST_MISSING_DETAILS", "CRITICAL", summary.blacklist_missing_details, "Blacklist missing date or reason"),
            ("Governance & Links", "Orphan Company Link", "ORPHAN_COMPANY_LINK", "CRITICAL", summary.company_orphan_links, "Link references non-existent company ID"),
            ("Governance & Links", "Duplicate Company Link", "DUPLICATE_COMPANY_LINKS", "WARNING", summary.company_duplicate_links, "Same company linked multiple times"),
            ("Governance & Links", "Company Missing Role", "COMPANY_MISSING_ROLE", "WARNING", summary.company_missing_role, "Company link missing role designation"),
            ("Governance & Links", "Orphan Extra Field", "EXTRA_FIELD_ORPHAN_ID", "CRITICAL", summary.extra_field_orphan_id, "Custom field with invalid schema definition"),
            ("Governance & Links", "Duplicate Extra Fields", "DUPLICATE_EXTRA_FIELDS", "WARNING", summary.extra_field_duplicate_entries, "Duplicate custom fields for Person"),
            ("Governance & Links", "Missing Deletion Date", "DELETED_MISSING_TIMESTAMP", "WARNING", summary.deleted_missing_del_date, "Deleted record missing deletion date"),
            ("Governance & Links", "Deletion Before Creation", "AUDIT_DEL_BEFORE_ENT", "CRITICAL", summary.audit_del_before_ent, "Deletion earlier than creation date"),
            ("Governance & Links", "Broken Zimbra Sync", "SYNC_ZIMBRA_MISSING_ID", "WARNING", summary.sync_zimbra_missing_id, "Sync enabled but missing Zimbra ID"),
        ]

        data_rows = [list(r) for r in rules_map]

        if is_csv:
            content = generate_csv(headers, data_rows)
        else:
            content = generate_xlsx("Quality Summary", headers, data_rows)

        return content, media_type, filename
